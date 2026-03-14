"""Excel document generator for creating structured workbook reports."""

from pathlib import Path
from typing import Dict, Any, List

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .base_generator import BaseDocumentGenerator, DocumentType


class ExcelGenerator(BaseDocumentGenerator):
    """Generator for Excel workbooks with structured metadata sheets.
    
    Creates multi-sheet Excel workbooks with formatted tables containing
    model metadata, authorship info, biological context, and references.
    """
    
    def __init__(self, *args, **kwargs):
        """Initialize the Excel generator.
        
        Raises:
            ImportError: If openpyxl is not installed
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError(
                "openpyxl is required for Excel generation. "
                "Install it with: pip install openpyxl"
            )
        super().__init__(*args, **kwargs)
    
    def get_file_extension(self) -> str:
        """Get the file extension for Excel files."""
        return ".xlsx"
    
    def _generate_impl(self,
                       output_path: Path,
                       document_data: Dict[str, Any],
                       document_type: DocumentType) -> None:
        """Generate an Excel workbook.
        
        Args:
            output_path: Where to save the Excel file
            document_data: All data for the document
            document_type: Type of document to generate
        """
        metadata = document_data["metadata"]
        
        # Create workbook
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet
        
        # Add sheets based on document type
        if document_type == DocumentType.TECHNICAL:
            # Consolidate metadata into fewer sheets
            self._add_summary_sheet(wb, metadata, document_data)
            self._add_consolidated_metadata_sheet(wb, metadata)  # Combines basic, authorship, bio context
            self._add_references_sheet(wb, metadata)
            
            # Add report panel data sheets if available (combined where possible)
            report_data = document_data.get("additional", {}).get("report_data", {})
            if report_data:
                # Combine model and topology into Analysis Results sheet
                has_model = 'model' in report_data and report_data['model'].get('has_data')
                has_topology = 'topology' in report_data and report_data['topology'].get('has_data')
                has_dynamic = 'dynamic' in report_data and report_data['dynamic'].get('has_data')
                
                if has_model or has_topology:
                    self._add_combined_analysis_sheet(wb, report_data)
                if has_dynamic:
                    self._add_dynamic_sheet(wb, report_data['dynamic'])
                if 'provenance' in report_data and report_data['provenance'].get('has_data'):
                    self._add_provenance_data_sheet(wb, report_data['provenance'])
        elif document_type == DocumentType.PUBLICATION:
            self._add_summary_sheet(wb, metadata, document_data)
            self._add_consolidated_metadata_sheet(wb, metadata)
            self._add_references_sheet(wb, metadata)
        else:  # SUMMARY
            self._add_summary_sheet(wb, metadata, document_data)
        
        # Save workbook
        wb.save(str(output_path))
    
    def _add_summary_sheet(self, wb: "Workbook", metadata: Dict[str, Any], 
                          document_data: Dict[str, Any]) -> None:
        """Add summary overview sheet."""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = metadata.get("basic", {}).get("model_name", "Untitled Model")
        ws['A1'].font = Font(size=16, bold=True, color="2C3E50")
        ws.merge_cells('A1:B1')
        
        # Generation info
        row = 2
        generation_date = document_data.get("generation_date")
        if generation_date:
            ws[f'A{row}'] = "Generated:"
            ws[f'B{row}'] = self.format_date(generation_date)
            row += 1
        
        row += 1  # Blank row
        
        # Key information
        self._add_header_row(ws, row, "Key Information")
        row += 1
        
        # Get authorship info
        authorship = metadata.get("authorship", {})
        primary_author = authorship.get("primary_author", "")
        contributors = authorship.get("contributors", [])
        author_display = primary_author
        if contributors:
            author_display += " et al."
        
        key_fields = [
            ("Model ID", metadata.get("basic", {}).get("model_id")),
            ("Version", metadata.get("basic", {}).get("version")),
            ("Organism", metadata.get("biological_context", {}).get("organism")),
            ("Authors", author_display if primary_author else ""),
            ("Institution", authorship.get("institution")),
        ]
        
        for label, value in key_fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
        
        # Description
        description = metadata.get("basic", {}).get("description")
        if description:
            row += 1
            ws[f'A{row}'] = "Description"
            self._style_header_cell(ws[f'A{row}'])
            row += 1
            ws[f'A{row}'] = description
            ws[f'A{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.merge_cells(f'A{row}:B{row}')
            ws.row_dimensions[row].height = 60
            row += 1
        
        # Report Data Summary (if available)
        report_data = document_data.get("additional", {}).get("report_data", {})
        if report_data:
            row += 1
            self._add_header_row(ws, row, "Report Summary")
            row += 1
            
            # Model structure summary
            model_data = report_data.get('model', {})
            if model_data and model_data.get('has_data'):
                structure = model_data.get('structure', {})
                ws[f'A{row}'] = "Model Structure"
                ws[f'B{row}'] = f"{structure.get('places_count', 0)} species, {structure.get('transitions_count', 0)} reactions, {structure.get('arcs_count', 0)} connections"
                self._style_data_row(ws, row)
                row += 1
            
            # Topology summary
            topology_data = report_data.get('topology', {})
            if topology_data and topology_data.get('has_data'):
                stats = topology_data.get('statistics', {})
                ws[f'A{row}'] = "Topology Analysis"
                ws[f'B{row}'] = f"{stats.get('p_invariants', 0)} P-invariants, {stats.get('t_invariants', 0)} T-invariants"
                self._style_data_row(ws, row)
                row += 1
            
            # Simulation summary
            dynamic_data = report_data.get('dynamic', {})
            if dynamic_data and dynamic_data.get('has_data'):
                sim_data = dynamic_data.get('simulation_data', {})
                species_count = len(sim_data.get('species', []))
                reactions_count = len(sim_data.get('reactions', []))
                if species_count > 0 or reactions_count > 0:
                    ws[f'A{row}'] = "Simulation Data"
                    summary_text = []
                    if species_count > 0:
                        summary_text.append(f"{species_count} species tracked")
                    if reactions_count > 0:
                        summary_text.append(f"{reactions_count} reactions analyzed")
                    ws[f'B{row}'] = ", ".join(summary_text)
                    self._style_data_row(ws, row)
                    row += 1
        
        self._auto_size_columns(ws)
    
    def _add_basic_info_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add basic information sheet."""
        ws = wb.create_sheet("Basic Information")
        basic = metadata.get("basic", {})
        
        row = 1
        self._add_header_row(ws, row, "Basic Information")
        row += 1
        
        fields = [
            ("Name", basic.get("model_name")),
            ("Model ID", basic.get("model_id")),
            ("Version", basic.get("version")),
            ("Description", basic.get("description")),
            ("Keywords", ", ".join(basic.get("keywords", [])) if basic.get("keywords") else None),
        ]
        
        for label, value in fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                if label == "Description":
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    ws.row_dimensions[row].height = 60
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_consolidated_metadata_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add consolidated metadata sheet combining basic info, authorship, and biological context."""
        ws = wb.create_sheet("Model Metadata")
        
        row = 1
        
        # Basic Information section
        basic = metadata.get("basic", {})
        self._add_header_row(ws, row, "Basic Information")
        row += 1
        
        basic_fields = [
            ("Name", basic.get("model_name")),
            ("Model ID", basic.get("model_id")),
            ("Version", basic.get("version")),
            ("Keywords", ", ".join(basic.get("keywords", [])) if basic.get("keywords") else None),
        ]
        
        for label, value in basic_fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
        
        # Description (with wrap)
        if basic.get("description"):
            ws[f'A{row}'] = "Description"
            ws[f'B{row}'] = basic.get("description")
            self._style_data_row(ws, row)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row].height = 60
            row += 1
        
        row += 1  # Blank row
        
        # Authorship section
        authorship = metadata.get("authorship", {})
        self._add_header_row(ws, row, "Authorship")
        row += 1
        
        authorship_fields = [
            ("Primary Author", authorship.get("primary_author")),
            ("Institution", authorship.get("institution")),
            ("Contributors", ", ".join(authorship.get("contributors", [])) if authorship.get("contributors") else None),
            ("Contact Email", authorship.get("contact_email")),
        ]
        
        for label, value in authorship_fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
        
        row += 1  # Blank row
        
        # Biological Context section
        bio = metadata.get("biological_context", {})
        self._add_header_row(ws, row, "Biological Context")
        row += 1
        
        bio_fields = [
            ("Organism", bio.get("organism")),
            ("Cellular Location", bio.get("cellular_location")),
            ("Biological Process", bio.get("biological_process")),
            ("Pathway", bio.get("pathway")),
            ("Disease Relevance", bio.get("disease_relevance")),
        ]
        
        for label, value in bio_fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
        
        row += 1  # Blank row
        
        # System Information section
        system = metadata.get("system", {})
        provenance = metadata.get("provenance", {})
        if system or provenance:
            self._add_header_row(ws, row, "System & Provenance")
            row += 1
            
            system_fields = [
                ("Created", system.get("created_date")),
                ("Last Modified", system.get("last_modified")),
                ("Shypn Version", system.get("shypn_version")),
                ("License", provenance.get("license")),
                ("Data Source", provenance.get("data_source")),
            ]
            
            for label, value in system_fields:
                if value:
                    ws[f'A{row}'] = label
                    if label in ["Created", "Last Modified"]:
                        ws[f'B{row}'] = self.format_date(value)
                    else:
                        ws[f'B{row}'] = self._safe_cell_value(value)
                    self._style_data_row(ws, row)
                    row += 1
        
        self._auto_size_columns(ws)
    
    def _add_combined_analysis_sheet(self, wb: "Workbook", report_data: Dict[str, Any]) -> None:
        """Add combined analysis sheet with model structure and topology data."""
        ws = wb.create_sheet("Analysis Results")
        
        row = 1
        self._add_header_row(ws, row, "Analysis Results")
        row += 1
        
        # Model Structure section
        model_data = report_data.get('model', {})
        if model_data and model_data.get('has_data'):
            self._add_header_row(ws, row, "Model Structure")
            row += 1
            
            structure = model_data.get('structure', {})
            for key, value in structure.items():
                label = key.replace('_', ' ').title()
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
            
            row += 1  # Blank row
        
        # Topology Analysis section
        topology_data = report_data.get('topology', {})
        if topology_data and topology_data.get('has_data'):
            self._add_header_row(ws, row, "Topology Analysis")
            row += 1
            
            # Status
            status = topology_data.get('status', 'No analysis')
            ws[f'A{row}'] = "Status"
            ws[f'B{row}'] = status
            self._style_data_row(ws, row)
            row += 1
            row += 1
            
            # Key findings
            findings = topology_data.get('key_findings', [])
            if findings:
                ws[f'A{row}'] = "Key Findings"
                ws[f'A{row}'].font = Font(bold=True, color="34495E")
                row += 1
                for idx, finding in enumerate(findings, 1):
                    ws[f'A{row}'] = f"  {idx}."
                    ws[f'B{row}'] = finding
                    self._style_data_row(ws, row)
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    row += 1
                row += 1
            
            # Consolidated sections
            sections = topology_data.get('sections', {})
            if sections:
                ws[f'A{row}'] = "Detailed Results"
                ws[f'A{row}'].font = Font(bold=True, color="34495E")
                row += 1
                
                for section_name, section_info in sections.items():
                    if section_info and isinstance(section_info, dict):
                        section_title = section_info.get('title', section_name.replace('_', ' ').title())
                        section_data = section_info.get('data', section_info)
                        
                        # Section category row
                        ws[f'A{row}'] = section_title
                        ws[f'A{row}'].font = Font(italic=True, color="7F8C8D")
                        row += 1
                        
                        # Section data
                        for key, value in section_data.items():
                            if key != 'title':
                                label = key.replace('_', ' ').title()
                                ws[f'A{row}'] = f"    {label}"
                                ws[f'B{row}'] = self._safe_cell_value(value)
                                self._style_data_row(ws, row)
                                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_basic_info_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add authorship information sheet."""
        ws = wb.create_sheet("Authorship")
        authorship = metadata.get("authorship", {})
        
        row = 1
        
        # Dates section
        self._add_header_row(ws, row, "Dates")
        row += 1
        
        # Get system dates
        system = metadata.get("system", {})
        date_fields = [
            ("Created", system.get("created_date")),
            ("Last Modified", system.get("last_modified")),
        ]
        
        for label, value in date_fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self.format_date(value)
                self._style_data_row(ws, row)
                row += 1
        
        row += 1  # Blank row
        
        # Author information section
        primary_author = authorship.get("primary_author")
        contributors = authorship.get("contributors", [])
        institution = authorship.get("institution")
        department = authorship.get("department")
        email = authorship.get("contact_email")
        
        if primary_author or contributors or institution:
            self._add_header_row(ws, row, "Author Information")
            row += 1
            
            if primary_author:
                ws[f'A{row}'] = "Primary Author"
                ws[f'B{row}'] = primary_author
                self._style_data_row(ws, row)
                row += 1
            
            if contributors:
                ws[f'A{row}'] = "Contributors"
                ws[f'B{row}'] = "; ".join(contributors)
                self._style_data_row(ws, row)
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = max(20, len(contributors) * 15)
                row += 1
            
            if institution:
                ws[f'A{row}'] = "Institution"
                ws[f'B{row}'] = institution
                self._style_data_row(ws, row)
                row += 1
            
            if department:
                ws[f'A{row}'] = "Department"
                ws[f'B{row}'] = department
                self._style_data_row(ws, row)
                row += 1
            
            if email:
                ws[f'A{row}'] = "Contact Email"
                ws[f'B{row}'] = email
                self._style_data_row(ws, row)
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_biological_context_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add biological context sheet."""
        ws = wb.create_sheet("Biological Context")
        bio_context = metadata.get("biological_context", {})
        
        row = 1
        self._add_header_row(ws, row, "Biological Context")
        row += 1
        
        fields = [
            ("Organism", bio_context.get("organism")),
            ("Biological System", bio_context.get("biological_system")),
            ("Pathway Name", bio_context.get("pathway_name")),
            ("Cell Type", bio_context.get("cell_type")),
        ]
        
        for label, value in fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                if label in ["Pathway Names", "GO Terms"]:
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    ws.row_dimensions[row].height = 40
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_provenance_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add provenance information sheet."""
        ws = wb.create_sheet("Provenance")
        provenance = metadata.get("provenance", {})
        
        row = 1
        self._add_header_row(ws, row, "Provenance")
        row += 1
        
        fields = [
            ("Source Database", provenance.get("source_database")),
            ("Database ID", provenance.get("database_id")),
            ("Original Format", provenance.get("original_format")),
            ("Import Date", self.format_date(provenance.get("import_date"))),
            ("Modifications", provenance.get("modifications")),
        ]
        
        for label, value in fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                if label == "Modifications":
                    ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                    ws.row_dimensions[row].height = 60
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_references_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add references sheet."""
        ws = wb.create_sheet("References")
        publications = metadata.get("references", {}).get("publications", [])
        
        row = 1
        self._add_header_row(ws, row, "References")
        row += 1
        
        if publications:
            # Column headers
            ws[f'A{row}'] = "#"
            ws[f'B{row}'] = "DOI / PubMed ID"
            for col in ['A', 'B']:
                self._style_header_cell(ws[f'{col}{row}'])
            row += 1
            
            # Publication data (simple string format)
            for idx, pub in enumerate(publications, 1):
                ws[f'A{row}'] = idx
                ws[f'B{row}'] = pub  # pub is just a string (DOI/PubMed ID)
                self._style_data_row(ws, row, columns=2)
                row += 1
        else:
            ws[f'A{row}'] = "No references available"
            row += 1
        
        self._auto_size_columns(ws)
    
    def _add_system_info_sheet(self, wb: "Workbook", metadata: Dict[str, Any]) -> None:
        """Add system information sheet."""
        ws = wb.create_sheet("System Information")
        system = metadata.get("system", {})
        
        row = 1
        self._add_header_row(ws, row, "System Information")
        row += 1
        
        fields = [
            ("File Format Version", system.get("file_format_version")),
        ]
        
        for label, value in fields:
            if value:
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_header_row(self, ws, row: int, title: str) -> None:
        """Add a styled header row."""
        ws[f'A{row}'] = title
        ws[f'A{row}'].font = Font(size=14, bold=True, color="FFFFFF")
        ws[f'A{row}'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells(f'A{row}:B{row}')
        ws.row_dimensions[row].height = 25
    
    def _safe_cell_value(self, value: Any) -> Any:
        """Convert value to Excel-compatible type.
        
        Args:
            value: The value to convert
            
        Returns:
            Excel-compatible value (converts lists/dicts to strings)
        """
        if isinstance(value, (list, dict, set, tuple)):
            if not value:  # Empty collection
                return "N/A"
            return str(value)
        return value
    
    def _style_header_cell(self, cell) -> None:
        """Apply header styling to a cell."""
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
        cell.alignment = Alignment(horizontal='left', vertical='center')
        thin_border = Border(
            left=Side(style='thin', color='AAAAAA'),
            right=Side(style='thin', color='AAAAAA'),
            top=Side(style='thin', color='AAAAAA'),
            bottom=Side(style='thin', color='AAAAAA')
        )
        cell.border = thin_border
    
    def _style_data_row(self, ws, row: int, columns: int = 2) -> None:
        """Apply data row styling."""
        thin_border = Border(
            left=Side(style='thin', color='DDDDDD'),
            right=Side(style='thin', color='DDDDDD'),
            top=Side(style='thin', color='DDDDDD'),
            bottom=Side(style='thin', color='DDDDDD')
        )
        
        for col_idx in range(columns):
            col_letter = get_column_letter(col_idx + 1)
            cell = ws[f'{col_letter}{row}']
            cell.border = thin_border
            
            if col_idx == 0:  # Label column
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
            
            cell.alignment = Alignment(horizontal='left', vertical='top')
    
    def _auto_size_columns(self, ws) -> None:
        """Auto-size columns based on content."""
        for column_cells in ws.columns:
            length = max(len(str(cell.value or "")) for cell in column_cells)
            col_letter = get_column_letter(column_cells[0].column)
            # Set width with some padding, max 60
            ws.column_dimensions[col_letter].width = min(length + 2, 60)
    
    def _format_authors_for_excel(self, primary_author: str, contributors: List[str]) -> str:
        """Format authors list for a single Excel cell.
        
        Args:
            primary_author: Primary author name
            contributors: List of contributor names
            
        Returns:
            Formatted author string
        """
        if not primary_author:
            return ""
        
        if contributors:
            return f"{primary_author}; " + "; ".join(contributors)
        else:
            return primary_author
    
    def _add_model_structure_sheet(self, wb: "Workbook", model_data: Dict[str, Any]) -> None:
        """Add model structure sheet with species and reactions."""
        ws = wb.create_sheet("Model Structure")
        
        row = 1
        
        # Overview section
        overview = model_data.get('overview', {})
        if overview:
            self._add_header_row(ws, row, "Overview")
            row += 1
            if overview.get('name'):
                ws[f'A{row}'] = "Model Name"
                ws[f'B{row}'] = overview['name']
                self._style_data_row(ws, row)
                row += 1
            if overview.get('description'):
                ws[f'A{row}'] = "Description"
                ws[f'B{row}'] = overview['description']
                self._style_data_row(ws, row)
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws.row_dimensions[row].height = 60
                row += 1
            row += 1
        
        # Structure counts
        structure = model_data.get('structure', {})
        if structure:
            self._add_header_row(ws, row, "Structure")
            row += 1
            ws[f'A{row}'] = "Species (Places)"
            ws[f'B{row}'] = structure.get('places_count', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "Reactions (Transitions)"
            ws[f'B{row}'] = structure.get('transitions_count', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "Connections (Arcs)"
            ws[f'B{row}'] = structure.get('arcs_count', 0)
            self._style_data_row(ws, row)
            row += 2
        
        # Species table
        species = model_data.get('species', [])
        if species:
            self._add_header_row(ws, row, "Species")
            row += 1
            
            # Column headers
            headers = ['ID', 'Name', 'Initial Tokens', 'Units', 'Mass', 'Conservation']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Species data
            for sp in species:
                ws[f'A{row}'] = sp.get('id', '')
                ws[f'B{row}'] = sp.get('name', '')
                ws[f'C{row}'] = sp.get('initial_tokens', '')
                ws[f'D{row}'] = sp.get('units', '')
                ws[f'E{row}'] = sp.get('mass', '')
                ws[f'F{row}'] = sp.get('conservation', '')
                self._style_data_row(ws, row, columns=6)
                row += 1
            row += 1
        
        # Reactions table
        reactions = model_data.get('reactions', [])
        if reactions:
            self._add_header_row(ws, row, "Reactions")
            row += 1
            
            # Column headers
            headers = ['ID', 'Name', 'Type', 'EC Number', 'Vmax', 'Km', 'Kcat', 'Ki', 'Rate Function', 'Reversible']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Reactions data
            for rxn in reactions:
                ws[f'A{row}'] = rxn.get('id', '')
                ws[f'B{row}'] = rxn.get('name', '')
                ws[f'C{row}'] = rxn.get('type', '')
                ws[f'D{row}'] = rxn.get('ec_number', '')
                ws[f'E{row}'] = rxn.get('vmax', '')
                ws[f'F{row}'] = rxn.get('km', '')
                ws[f'G{row}'] = rxn.get('kcat', '')
                ws[f'H{row}'] = rxn.get('ki', '')
                ws[f'I{row}'] = rxn.get('rate_function', '')
                ws[f'J{row}'] = rxn.get('reversible', '')
                self._style_data_row(ws, row, columns=10)
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_topology_sheet(self, wb: "Workbook", topology_data: Dict[str, Any]) -> None:
        """Add topology analyses sheet."""
        ws = wb.create_sheet("Topology Analyses")
        
        row = 1
        self._add_header_row(ws, row, "Topology Analyses")
        row += 1
        
        # Status
        status = topology_data.get('status', 'No analysis')
        ws[f'A{row}'] = "Status"
        ws[f'B{row}'] = status
        self._style_data_row(ws, row)
        row += 2
        
        # Statistics
        statistics = topology_data.get('statistics', {})
        if statistics:
            self._add_header_row(ws, row, "Statistics")
            row += 1
            for key, value in statistics.items():
                label = key.replace('_', ' ').title()
                ws[f'A{row}'] = label
                ws[f'B{row}'] = self._safe_cell_value(value)
                self._style_data_row(ws, row)
                row += 1
            row += 1
        
        # Key findings
        findings = topology_data.get('key_findings', [])
        if findings:
            self._add_header_row(ws, row, "Key Findings")
            row += 1
            for idx, finding in enumerate(findings, 1):
                ws[f'A{row}'] = f"{idx}."
                ws[f'B{row}'] = finding
                self._style_data_row(ws, row)
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                row += 1
            row += 1
        
        # Warnings
        warnings = topology_data.get('warnings', [])
        if warnings:
            self._add_header_row(ws, row, "Warnings")
            row += 1
            for idx, warning in enumerate(warnings, 1):
                ws[f'A{row}'] = f"{idx}."
                ws[f'B{row}'] = warning
                self._style_data_row(ws, row)
                ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
                ws[f'B{row}'].font = Font(color="E74C3C")
                row += 1
            row += 1
        
        # Consolidated Analysis sections - all in one compact table
        sections = topology_data.get('sections', {})
        if sections:
            self._add_header_row(ws, row, "Detailed Analyses")
            row += 1
            
            # Consolidate all section data into single table
            for section_name, section_info in sections.items():
                if section_info and isinstance(section_info, dict):
                    section_title = section_info.get('title', section_name.replace('_', ' ').title())
                    section_data = section_info.get('data', section_info)
                    
                    # Section header as a category row
                    ws[f'A{row}'] = section_title
                    ws[f'A{row}'].font = Font(bold=True, color="34495E")
                    ws.merge_cells(f'A{row}:B{row}')
                    row += 1
                    
                    # Add section data compactly
                    for key, value in section_data.items():
                        if key != 'title':  # Skip title if it's in data dict
                            label = key.replace('_', ' ').title()
                            ws[f'A{row}'] = f"  {label}"  # Indent for visual grouping
                            ws[f'B{row}'] = self._safe_cell_value(value)
                            self._style_data_row(ws, row)
                            row += 1
        
        self._auto_size_columns(ws)
    
    def _add_dynamic_sheet(self, wb: "Workbook", dynamic_data: Dict[str, Any]) -> None:
        """Add dynamic analyses sheet with simulation data (Species Concentration & Reaction Activity)."""
        ws = wb.create_sheet("Dynamic Analyses")
        
        row = 1
        self._add_header_row(ws, row, "Dynamic Analyses")
        row += 1
        
        # Simulation Parameters (NEW - appears first)
        sim_params = dynamic_data.get('simulation_parameters', {})
        if sim_params:
            self._add_header_row(ws, row, "Simulation Parameters")
            row += 1
            
            # Timestamp
            timestamp = sim_params.get('timestamp')
            if timestamp:
                ws[f'A{row}'] = "Date/Time"
                ws[f'B{row}'] = str(timestamp)
                self._style_data_row(ws, row)
                row += 1
            
            # Time parameters
            time_step = sim_params.get('time_step')
            if time_step is not None:
                ws[f'A{row}'] = "Time Step (dt)"
                ws[f'B{row}'] = f"{time_step:.4f} s"
                self._style_data_row(ws, row)
                row += 1
            
            target_duration = sim_params.get('target_duration')
            if target_duration is not None:
                ws[f'A{row}'] = "Target Duration"
                ws[f'B{row}'] = f"{target_duration:.2f} s"
                self._style_data_row(ws, row)
                row += 1
            
            actual_duration = sim_params.get('actual_duration', 0)
            ws[f'A{row}'] = "Actual Duration"
            ws[f'B{row}'] = f"{actual_duration:.2f} s"
            self._style_data_row(ws, row)
            row += 1
            
            time_scale = sim_params.get('time_scale', 1.0)
            ws[f'A{row}'] = "Time Scale"
            ws[f'B{row}'] = f"{time_scale:.1f}x"
            self._style_data_row(ws, row)
            row += 1
            
            total_steps = sim_params.get('total_steps', 0)
            ws[f'A{row}'] = "Total Steps"
            ws[f'B{row}'] = total_steps
            self._style_data_row(ws, row)
            row += 1
            
            # Activity statistics
            num_time_points = sim_params.get('num_time_points', 0)
            ws[f'A{row}'] = "Time Points Recorded"
            ws[f'B{row}'] = num_time_points
            self._style_data_row(ws, row)
            row += 1
            
            total_firings = sim_params.get('total_firings', 0)
            ws[f'A{row}'] = "Total Transition Firings"
            ws[f'B{row}'] = total_firings
            self._style_data_row(ws, row)
            row += 1
            
            avg_rate = sim_params.get('avg_firing_rate', 0)
            ws[f'A{row}'] = "Average Firing Rate"
            ws[f'B{row}'] = f"{avg_rate:.2f} firings/s"
            self._style_data_row(ws, row)
            row += 2
        
        # Model info
        model_info = dynamic_data.get('model_info', {})
        if model_info:
            self._add_header_row(ws, row, "Model Information")
            row += 1
            ws[f'A{row}'] = "Number of Species"
            ws[f'B{row}'] = model_info.get('num_places', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "Number of Reactions"
            ws[f'B{row}'] = model_info.get('num_transitions', 0)
            self._style_data_row(ws, row)
            row += 2
        
        # Organisms
        organisms = dynamic_data.get('organisms', [])
        if organisms:
            ws[f'A{row}'] = "Organisms"
            ws[f'B{row}'] = "; ".join(organisms)
            self._style_data_row(ws, row)
            ws[f'B{row}'].alignment = Alignment(wrap_text=True, vertical='top')
            row += 1
        
        # Enrichments
        enrichments_count = dynamic_data.get('enrichments_count', 0)
        if enrichments_count > 0:
            ws[f'A{row}'] = "Enrichments"
            ws[f'B{row}'] = f"{enrichments_count} parameter(s) enriched"
            self._style_data_row(ws, row)
            row += 2
        else:
            row += 1
        
        # Species Concentration Table (from simulation data)
        simulation = dynamic_data.get('simulation_data', {})
        species_sim = simulation.get('species', [])
        if species_sim:
            self._add_header_row(ws, row, "Species Concentration")
            row += 1
            
            # Column headers
            headers = ['Name', 'ID', 'Initial', 'Final', 'Min', 'Max', 'Avg']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Species concentration data
            for sp in species_sim:
                ws[f'A{row}'] = sp.get('name', '')
                ws[f'B{row}'] = sp.get('id', '')
                ws[f'C{row}'] = sp.get('initial', '')
                ws[f'D{row}'] = sp.get('final', '')
                ws[f'E{row}'] = sp.get('min', '')
                ws[f'F{row}'] = sp.get('max', '')
                ws[f'G{row}'] = sp.get('avg', '')
                self._style_data_row(ws, row, columns=7)
                row += 1
            row += 1
        
        # Reaction Activity Table (from simulation data)
        reactions_sim = simulation.get('reactions', [])
        if reactions_sim:
            self._add_header_row(ws, row, "Reaction Activity")
            row += 1
            
            # Column headers
            headers = ['Name', 'ID', 'Avg Rate', 'Total Firings', 'Status']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Reaction activity data
            for rxn in reactions_sim:
                ws[f'A{row}'] = rxn.get('name', '')
                ws[f'B{row}'] = rxn.get('id', '')
                ws[f'C{row}'] = rxn.get('avg_rate', '')
                ws[f'D{row}'] = rxn.get('total_firings', '')
                ws[f'E{row}'] = rxn.get('status', '')
                self._style_data_row(ws, row, columns=5)
                row += 1
            row += 1
        
        # Reaction Selected Table (Locality simulation data)
        reactions_selected = simulation.get('reactions_selected', [])
        if reactions_selected:
            self._add_header_row(ws, row, "Reaction Selected (Locality)")
            row += 1
            
            # Column headers
            headers = ['Component', 'ID', 'Name', 'Initial', 'Final', 'Min', 'Max', 'Avg', 'Info']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Reaction selected data
            for rxn in reactions_selected:
                ws[f'A{row}'] = rxn.get('component', '')
                ws[f'B{row}'] = rxn.get('id', '')
                ws[f'C{row}'] = rxn.get('name', '')
                ws[f'D{row}'] = rxn.get('initial', '')
                ws[f'E{row}'] = rxn.get('final', '')
                ws[f'F{row}'] = rxn.get('min', '')
                ws[f'G{row}'] = rxn.get('max', '')
                ws[f'H{row}'] = rxn.get('average', '')
                ws[f'I{row}'] = rxn.get('info', '')
                self._style_data_row(ws, row, columns=9)
                row += 1
        
        self._auto_size_columns(ws)
    
    def _add_provenance_data_sheet(self, wb: "Workbook", provenance_data: Dict[str, Any]) -> None:
        """Add provenance data sheet with pathways information."""
        ws = wb.create_sheet("Data Provenance")
        
        row = 1
        self._add_header_row(ws, row, "Data Provenance")
        row += 1
        
        # Summary
        summary = provenance_data.get('summary', {})
        if summary:
            ws[f'A{row}'] = "Total Pathways"
            ws[f'B{row}'] = summary.get('total', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "KEGG Pathways"
            ws[f'B{row}'] = summary.get('kegg', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "SBML Models"
            ws[f'B{row}'] = summary.get('sbml', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "Converted"
            ws[f'B{row}'] = summary.get('converted', 0)
            self._style_data_row(ws, row)
            row += 1
            ws[f'A{row}'] = "Enriched"
            ws[f'B{row}'] = summary.get('enriched', 0)
            self._style_data_row(ws, row)
            row += 2
        
        # Pathways table
        pathways = provenance_data.get('pathways', [])
        if pathways:
            self._add_header_row(ws, row, "Pathways")
            row += 1
            
            # Column headers
            headers = ['Name', 'Source', 'Source ID', 'Organism', 'Import Date', 'Enrichments']
            for col_idx, header in enumerate(headers):
                col_letter = get_column_letter(col_idx + 1)
                ws[f'{col_letter}{row}'] = header
                self._style_header_cell(ws[f'{col_letter}{row}'])
            row += 1
            
            # Pathways data
            for pathway in pathways:
                ws[f'A{row}'] = pathway.get('name', '')
                ws[f'B{row}'] = pathway.get('source_type', '')
                ws[f'C{row}'] = pathway.get('source_id', '')
                ws[f'D{row}'] = pathway.get('organism', '')
                ws[f'E{row}'] = pathway.get('import_date', '')
                ws[f'F{row}'] = pathway.get('enrichments_count', 0)
                self._style_data_row(ws, row, columns=6)
                row += 1
        
        self._auto_size_columns(ws)
